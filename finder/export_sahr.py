import openpyxl
import pandas as pd
from sqlalchemy import create_engine

from .models import Data_Table


def doc_sahr():
    wb = openpyxl.load_workbook("/home/banywick/Projects/enterprise/Сахр актуальное деплой.xlsx")
    sheet = wb['Лист1']
    for row in sheet.iter_rows(min_row=0, max_row=6106, min_col=0, max_col=0, values_only=True):
        Data_Table.objects.create(
            article=row[0],
            title=row[1],
            address=row[2],
            comment=row[3],
            date=row[4],
        )
# doc_sahr()            
# def doc_sahr():
#     wb = openpyxl.load_workbook("/home/banywick/Projects/enterprise/образец данных.xlsx")
#     sheet = wb['Лист1']
#     for row in sheet.iter_rows(min_row=0, max_row=6106, min_col=0, max_col=0, values_only=True):
#         print( row[4])
# # doc_sahr()            
